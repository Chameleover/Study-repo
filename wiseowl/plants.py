# importing openpyxl module
import openpyxl

# Give the location of the file
path = "/Users/andonov/Documents/GitHub/Python Course/wiseowl/Plants.xlsx"

# workbook object is created
wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active
max_col = sheet_obj.max_column

# Loop will print all columns name
for i in range(1, max_col + 1):
    in_stock = sheet_obj.cell(row=i, column=8)
    obj_name = sheet_obj.cell(row=i, column=1)
    # print(cell_obj.value)

    # Check if in stock

    if in_stock.value == 'No':
        print(obj_name.value)

print('\nThe above plants are not in stock')